from openpyxl import load_workbook
from features.notas_fiscais.model import NotaFiscal
from openpyxl.utils.cell import (
    get_column_letter,
    range_boundaries
)

NOME_ABA = "Conferência"
PRIMEIRA_LINHA_DADOS = 6

COLUNA_NF = 2
COLUNA_VALOR = 3
COLUNA_EMISSAO = 4
COLUNA_CHAVE = 7
COLUNA_FORNECEDOR = 8
COLUNA_CNPJ = 9
COLUNA_MATERIAIS = 10

class NotaFiscalRepository:

    def salvar(self, caminho_planilha: str, notas: list[NotaFiscal]) -> None:
        workbook = load_workbook(caminho_planilha)
        sheet = workbook[NOME_ABA]

        for nota in notas:
            proxima_linha = self._buscar_proxima_linha(sheet)

            sheet.cell(proxima_linha, COLUNA_NF, nota.numero)
            sheet.cell(proxima_linha, COLUNA_VALOR, nota.valor)
            sheet.cell(proxima_linha, COLUNA_EMISSAO, nota.emissao)
            sheet.cell(proxima_linha, COLUNA_CHAVE, nota.chave)
            sheet.cell(proxima_linha, COLUNA_FORNECEDOR, nota.fornecedor)
            sheet.cell(proxima_linha, COLUNA_CNPJ, nota.cnpj)
            sheet.cell(proxima_linha, COLUNA_MATERIAIS, "\n".join(nota.materiais))

        ultima_linha = self._buscar_proxima_linha(sheet) - 1

        self._estender_tabela(
            sheet,
            ultima_linha
        )

        workbook.save(caminho_planilha)
        workbook.close()

    def _buscar_proxima_linha(self, sheet) -> int:
        linha = PRIMEIRA_LINHA_DADOS

        while sheet.cell(linha, COLUNA_NF).value is not None:
            linha += 1

        return linha

    def buscar_chaves_existentes(self,caminho_planilha: str) -> set[str]:
        workbook = load_workbook(caminho_planilha)
        sheet = workbook[NOME_ABA]

        chaves_existentes = set()

        for linha in range(PRIMEIRA_LINHA_DADOS, sheet.max_row + 1):
            chave = sheet.cell(linha, COLUNA_CHAVE).value

            if chave is None:
                continue

            chave_normalizada = self._normalizar_chave(chave)

            if chave_normalizada:
                chaves_existentes.add(chave_normalizada)

        workbook.close()

        return chaves_existentes

    def _estender_tabela(self, sheet, ultima_linha: int) -> None:
        for tabela in sheet.tables.values():
            coluna_inicial, linha_inicial, coluna_final, linha_final = (
                range_boundaries(tabela.ref)
            )

            if ultima_linha <= linha_final:
                continue

            tabela.ref = (
                f"{get_column_letter(coluna_inicial)}"
                f"{linha_inicial}:"
                f"{get_column_letter(coluna_final)}"
                f"{ultima_linha}"
            )

    @staticmethod
    def _normalizar_chave(chave) -> str:
        if chave is None:
            return ""

        return "".join(
            caractere
            for caractere in str(chave)
            if caractere.isdigit()
        )